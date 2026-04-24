from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from .config_reader import DATE_FORMAT, LoadedConfig


RAG_FILLS = {
    "GREEN": PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),
    "AMBER": PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"),
    "RED": PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
}


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)


def _apply_basic_sheet_formatting(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def _billing_status(row: pd.Series) -> str:
    deficit = row.get("deficit_at_risk")
    surplus = row.get("surplus_cost_bearing")
    if pd.notna(deficit) and float(deficit) > 0:
        return f"Deficit {float(deficit):.2f} hrs"
    if pd.notna(surplus) and float(surplus) > 0:
        return f"Surplus {float(surplus):.2f} hrs"
    return "On target"


def generate_report(
    fte_df: pd.DataFrame,
    accuracy_df: pd.DataFrame,
    productivity_df: pd.DataFrame,
    ur_df: pd.DataFrame,
    billing_df: pd.DataFrame,
    master_df: pd.DataFrame,
    onduty_filtered: pd.DataFrame,
    productivity_filtered: pd.DataFrame,
    ur_filtered: pd.DataFrame,
    accuracy_filtered: pd.DataFrame,
    loaded_config: LoadedConfig,
) -> Path:
    output_folder = loaded_config.run_config.output_folder
    output_folder.mkdir(parents=True, exist_ok=True)

    filename = f"SENTINEL_Report_{loaded_config.run_config.end_date.strftime(DATE_FORMAT)}.xlsx"
    output_path = output_folder / filename

    dashboard = (
        fte_df[["project_name", "achievement_pct", "rag_status"]]
        .rename(columns={"achievement_pct": "FTE%", "rag_status": "FTE RAG"})
        .merge(
            accuracy_df[["project_name", "achievement_pct", "rag_status"]].rename(
                columns={"achievement_pct": "Accuracy%", "rag_status": "Accuracy RAG"}
            ),
            on="project_name",
            how="outer",
        )
        .merge(
            productivity_df[["project_name", "achievement_pct", "rag_status"]].rename(
                columns={"achievement_pct": "Productivity%", "rag_status": "Productivity RAG"}
            ),
            on="project_name",
            how="outer",
        )
        .merge(
            ur_df[["project_name", "achievement_pct", "rag_status"]].rename(
                columns={"achievement_pct": "UR%", "rag_status": "UR RAG"}
            ),
            on="project_name",
            how="outer",
        )
        .merge(
            billing_df[["project_name", "surplus_cost_bearing", "deficit_at_risk", "rag_status"]].rename(
                columns={"rag_status": "Billing RAG"}
            ),
            on="project_name",
            how="outer",
        )
    )
    dashboard["Billing Status"] = dashboard.apply(_billing_status, axis=1)
    dashboard = dashboard.rename(columns={"project_name": "Project"})
    dashboard = dashboard[
        [
            "Project",
            "FTE%",
            "FTE RAG",
            "Accuracy%",
            "Accuracy RAG",
            "Productivity%",
            "Productivity RAG",
            "UR%",
            "UR RAG",
            "Billing Status",
            "Billing RAG",
        ]
    ].sort_values("Project", ignore_index=True)

    training_df = master_df.loc[master_df.get("non_billable", False).fillna(False), ["email", "project_name"]].copy()
    training_df = training_df.rename(columns={"email": "labeller", "project_name": "Project"})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, "Dashboard", dashboard)
        _write_sheet(writer, "FTE Analysis", fte_df)
        _write_sheet(writer, "Accuracy", accuracy_df)
        _write_sheet(writer, "Productivity", productivity_df)
        _write_sheet(writer, "UR Analysis", ur_df)
        _write_sheet(writer, "Billing Forecast", billing_df)
        _write_sheet(writer, "Training Labellers", training_df)
        _write_sheet(writer, "Raw Combined", master_df)
        _write_sheet(writer, "OnDuty Raw", onduty_filtered)
        _write_sheet(writer, "Productivity Raw", productivity_filtered)
        _write_sheet(writer, "UR Raw", ur_filtered)
        _write_sheet(writer, "Accuracy Raw", accuracy_filtered)

        for worksheet in writer.book.worksheets:
            _apply_basic_sheet_formatting(worksheet)

        dashboard_sheet = writer.book["Dashboard"]
        rag_columns = ["C", "E", "G", "I", "K"]
        for column_letter in rag_columns:
            for cell in dashboard_sheet[column_letter][1:]:
                fill = RAG_FILLS.get(cell.value)
                if fill is not None:
                    cell.fill = fill

    print(f"[REPORT] Excel report created: {output_path}")
    return output_path
